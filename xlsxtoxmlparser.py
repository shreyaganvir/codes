"""
Python program to read given xls file and convert data to xml

@author Shreyali Ganvir
Date    30th June 2024

"""

import argparse
from openpyxl import load_workbook
import xml.etree.ElementTree as et
from xml.dom import minidom


def create_xml_from_xlsx(file_path):
    # Load the workbook
    workbook = load_workbook(file_path)

    # Select the first sheet
    sheet = workbook.active

    # Skip the header row
    data_rows = sheet.iter_rows(min_row=2, values_only=True)

    # Define profiles (assuming 'Summer' and 'Winter' as per your example)
    profiles = ['Summer', 'Winter']

    # Iterate through the rows and columns of xlsx tp parse into xml
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract data from the row
        alias = row[0]
        limits_data = row[1:]
        # Initialize the XML structure
        limits = et.Element('limits')
        newlimitset = et.SubElement(limits, 'newlimitset', template="Voltage (Volts)", attribute_name="Limitband",
                                    created_by="SCT", alias=alias, hireason="32")
        # Create limittype elements
        for limit_type, limit_value in zip(['LoLo', 'Lo', 'Hi', 'HiHi'], limits_data):
            limittype = et.SubElement(newlimitset, 'limittype', type=limit_type, hysteresis_type="Absolute",
                                      hysteresis_param="0")

        # Create profile elements
        for profile in profiles:
            profile_elem = et.SubElement(newlimitset, 'profile', name=profile, nominal="0")

            # Add limitvalue elements
            for limit_type, limit_value in zip(['LoLo', 'Lo', 'Hi', 'HiHi'], limits_data):
                limitvalue = et.SubElement(profile_elem, 'limitvalue', type=limit_type, value_type="Absolute",
                                           value_param=str(limit_value))

        # Pretty-print the XML
        xml_str = et.tostring(limits, encoding='utf-8')
        parsed_xml = minidom.parseString(xml_str)
        pretty_xml_str = parsed_xml.toprettyxml(indent="  ")

        # Write the XML to file with headers on new lines
        with open('07-LIM_{}.xml'.format(alias), 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<!DOCTYPE limits SYSTEM "limits.dtd">\n')
            f.write(pretty_xml_str[parsed_xml.toprettyxml(indent="  ").find('\n') + 1:])

        print("07-LIM_{}.xml. XML file generated successfully.".format(alias))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Excel file to XML format.")
    parser.add_argument("--input_file", help="Path to the input Excel file.", required=True)
    args = parser.parse_args()
    create_xml_from_xlsx(args.input_file)
